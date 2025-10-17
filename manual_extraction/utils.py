import json, os, random, re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment
from openpyxl.workbook.defined_name import DefinedName

import shutil
from pathlib import Path


def make_named_range_key(s: str, used: set) -> str:
    """Excel-safe, unique name for a named range (workbook-scoped)."""
    base = re.sub(r"[^A-Za-z0-9_]", "_", s.strip())
    if not base or not re.match(r"^[A-Za-z_]", base):
        base = f"Q_{base}"
    base = base[:240]  # leave room for collision suffix
    key = base
    i = 1
    while key in used:
        key = f"{base}_{i}"
        i += 1
    used.add(key)
    return key


def upsert_defined_name(wb, name: str, ref: str) -> None:
    """
    Remove any existing workbook-defined names whose .name matches `name`
    (case-insensitive), then add a fresh name -> ref.
    Compatible with openpyxl 3.1.x (DefinedNameDict).
    """
    for key in list(wb.defined_names):  # iterate keys to safely delete
        dn = wb.defined_names[key]
        if getattr(dn, "name", "").lower() == name.lower():
            del wb.defined_names[key]
    wb.defined_names.add(DefinedName(name=name, attr_text=ref))


def build_workbook(
    papers,
    questions_dict,
    out_path="dropdown_menus.xlsm",
    start_row=2,
    max_rows=200,
    template_path="excel_macro.xlsm",
):
    """
    Build an .xlsm workbook using a macro-enabled template that contains
    a Worksheet_Change handler in the 'Papers' sheet to support multi-select.
    """
    # Load macro-enabled template so VBA is preserved
    wb = load_workbook(template_path, keep_vba=True)

    # --- lists sheet (source values for validation) ---
    if "lists" in wb.sheetnames:
        ws_lists = wb["lists"]
        ws_lists.delete_rows(1, ws_lists.max_row or 1)
    else:
        ws_lists = wb.create_sheet("lists")

    used_names = set()
    name_for_question = {}

    for col_idx, (q, opts) in enumerate(questions_dict.items(), start=1):
        ws_lists.cell(row=1, column=col_idx, value=q)
        opts = ["" if v is None else str(v) for v in opts]
        for r, opt in enumerate(opts, start=2):
            ws_lists.cell(row=r, column=col_idx, value=opt)

        last_row = 1 + max(len(opts), 1)
        col_letter = get_column_letter(col_idx)
        ref = f"lists!${col_letter}$2:${col_letter}${last_row}"

        name = make_named_range_key(q, used_names)
        upsert_defined_name(wb, name, ref)
        name_for_question[q] = name

    # --- main sheet (Papers) ---
    if "Papers" in wb.sheetnames:
        ws_main = wb["Papers"]
        ws_main.delete_rows(1, ws_main.max_row or 1)
    else:
        ws_main = wb.create_sheet("Papers")

    # Header row
    ws_main.cell(row=1, column=1, value="Paper")
    for j, q in enumerate(questions_dict.keys(), start=2):
        c = ws_main.cell(row=1, column=j, value=q)
        c.alignment = Alignment(wrap_text=True)

    # Paper names
    for i, p in enumerate(papers, start=start_row):
        ws_main.cell(row=i, column=1, value=p)

    # Data validations (point at workbook-defined names created above)
    num_cols = 1 + len(questions_dict)
    for j, q in enumerate(questions_dict.keys(), start=2):
        named = name_for_question[q]
        dv = DataValidation(type="list", formula1=f"={named}", allow_blank=True)
        ws_main.add_data_validation(dv)
        dv.add(
            f"{get_column_letter(j)}{start_row}:"
            f"{get_column_letter(j)}{start_row + max_rows - 1}"
        )

    # Basic formatting
    ws_main.freeze_panes = "B2"
    ws_main.column_dimensions["A"].width = 40
    for j in range(2, num_cols + 1):
        ws_main.column_dimensions[get_column_letter(j)].width = 32

    # Save as macro-enabled workbook
    wb.save(out_path)
    return out_path
